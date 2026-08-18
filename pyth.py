import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

input_file = "failuresClassifications_MGU22_03-08-2026-10-08-2026.xlsx"
output_file = "Irregular_Test_Failures.xlsx"

# Delete existing output file
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except PermissionError:
        print("Please close the output Excel file and try again.")
        exit()

# Read Report sheet
temp_df = pd.read_excel(
    input_file,
    sheet_name="Report",
    header=None
)

# Find header row
header_row = None

for i in range(len(temp_df)):
    row_values = temp_df.iloc[i].astype(str).tolist()

    if "CATEGORY" in row_values:
        header_row = i
        break

# Read actual data
df = pd.read_excel(
    input_file,
    sheet_name="Report",
    header=header_row
)

# Clean column names
df.columns = df.columns.astype(str).str.strip()

# Filter only irregular failures
filtered_df = df[
    df["CATEGORY"].astype(str).str.strip()
    == "IRREGULAR_TEST_FAILURE"
]

# Replace empty Ticket values with NULL
if "TICKET" in filtered_df.columns:
    filtered_df["TICKET"] = (
        filtered_df["TICKET"]
        .fillna("NULL")
        .replace("", "NULL")
    )

# Save to Excel
filtered_df.to_excel(output_file, index=False)

# Open workbook for formatting
wb = load_workbook(output_file)
ws = wb.active

# Rename sheet
ws.title = "Report"

# Header formatting
for cell in ws[1]:
    cell.font = Font(
        bold=True,
        color="000000"
    )

    cell.fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

# Freeze first row
ws.freeze_panes = "A2"

# Wrap text
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

# Auto adjust column widths
for column in ws.columns:

    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
        except:
            pass

    ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

# Increase row height
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 35

# Zoom
ws.sheet_view.zoomScale = 85

# Save workbook
wb.save(output_file)

print("Completed Successfully")
print("Output File:", output_file)
print("Total Irregular Failures:", len(filtered_df))
